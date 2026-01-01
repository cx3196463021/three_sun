from openpyxl import load_workbook
import os
import re
import requests
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from tenacity import retry, stop_after_attempt, wait_random
import json
from datetime import datetime
import sys
import asyncio
import aiohttp
from aiohttp import ClientSession, TCPConnector
thread_local = threading.local()

def get_resource_path(relative_path):
    """获取资源文件的绝对路径（用于打包进exe的资源）"""
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def get_data_path(relative_path):
    """获取外部数据文件的绝对路径（用户自己放的数据）"""
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def get_history_data_folder():
    """获取历史数据文件夹路径，如果不存在则创建"""
    folder_path = get_data_path('历史数据文件夹')
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

def get_error_data_folder():
    """获取错误数据文件夹路径，如果不存在则创建"""
    folder_path = get_data_path('错误数据')
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

def get_history_data_filename(date_str=None):
    """生成历史数据文件名
    Args:
        date_str: 日期字符串，格式 '2025-10-30'，如果为None则使用今天
    Returns:
        str: 文件名，如 '10-30-东方财富历史数据.json'
    """
    if date_str is None:
        now = datetime.now()
        month = now.month
        day = now.day
    else:
        parts = date_str.split('-')
        month = int(parts[1])
        day = int(parts[2])
    return f'{month}-{day}-东方财富历史数据.json'

def save_failed_stocks(failed_stocks, date_str=None, prefix='history'):
    """保存采集失败的股票代码到txt文件
    Args:
        failed_stocks: 失败的股票代码列表
        date_str: 日期字符串，如果为None则使用今天
        prefix: 文件名前缀，'history'（历史数据）或'realtime'（实时数据）
    """
    if not failed_stocks:
        return
    folder_path = get_error_data_folder()
    if date_str is None:
        now = datetime.now()
        filename = f'{now.month}-{now.day}-{prefix}-失败代码.txt'
    else:
        parts = date_str.split('-')
        month = int(parts[1])
        day = int(parts[2])
        filename = f'{month}-{day}-{prefix}-失败代码.txt'
    file_path = os.path.join(folder_path, filename)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f'采集失败的股票代码 ({len(failed_stocks)}个)\n')
        f.write(f"采集时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write('==================================================\n')
        for stock_code in failed_stocks:
            f.write(f'{stock_code}\n')
    print(f'采集失败的股票代码已保存到: {file_path}')

def save_history_data_to_file(history_data, date_str=None):
    """保存历史数据到JSON文件
    Args:
        history_data: 历史数据字典，格式 {代码: {历史价格列表: [...], ...}}
        date_str: 日期字符串，如果为None则使用今天
    """
    folder_path = get_history_data_folder()
    filename = get_history_data_filename(date_str)
    file_path = os.path.join(folder_path, filename)
    formatted_data = {}
    for stock_code, stock_data in history_data.items():
        price_list = stock_data.get('历史价格列表', [])
        price_list_sorted = sorted(price_list, key=lambda x: x['日期'])
        formatted_list = []
        for price_data in price_list_sorted:
            formatted_list.append({'日期': price_data['日期'], '收盘价': price_data['收盘价'], '涨幅': price_data['涨幅']})
        formatted_data[stock_code] = {'代码': stock_code, '历史价格列表': formatted_list, '昨日收盘价': stock_data.get('昨日收盘价', 0), '昨日涨幅': stock_data.get('昨日涨幅', 0), '30日最高价': stock_data.get('30日最高价', 0), '30日最低价': stock_data.get('30日最低价', 0), '60日最高价': stock_data.get('60日最高价', 0), '60日最低价': stock_data.get('60日最低价', 0)}
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(formatted_data, f, ensure_ascii=False, indent=2)
    print(f'历史数据已保存到: {file_path} (共{len(formatted_data)}个股票)')

def load_history_data_from_file(date_str=None):
    """从JSON文件读取历史数据
    Args:
        date_str: 日期字符串，如果为None则使用今天
    Returns:
        dict: 历史数据，如果文件不存在则返回None
    """
    folder_path = get_history_data_folder()
    filename = get_history_data_filename(date_str)
    file_path = os.path.join(folder_path, filename)
    if not os.path.exists(file_path):
        return
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f'从文件读取历史数据: {file_path} (共{len(data)}个股票)')
        return data
    except Exception as e:
        print(f'读取历史数据文件失败: {e}')
        return None

def get_session():
    """获取当前线程的 session"""
    if not hasattr(thread_local, 'session'):
        thread_local.session = requests.Session()
        thread_local.session.headers.update({'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
    return thread_local.session

def get_folder_data(strat_index=1, count=1):
    xlsx_datas = {}
    file_name = []
    folder_name = '股票数据'
    base_path = get_data_path('')
    stock_path = os.path.join(base_path, folder_name)
    files = os.listdir(stock_path)
    for file in files:
        if file.startswith('~$'):
            continue
        if not file.endswith('.xlsx'):
            continue
        file_name.append(file)
    reverse_files = sorted(file_name, reverse=True)
    result_files = reverse_files[strat_index:strat_index + count]
    for data_file in result_files:
        file_path = os.path.join(stock_path, data_file)
        try:
            wb = load_workbook(file_path)
        except Exception as e:
            print(f'警告：无法读取文件 {data_file}，错误：{e}')
            continue
        ws = wb.active
        rows_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3]:
                concept = row[3].split('+')[0]
            row_data = [row[0], row[1], row[2], concept, row[4], row[5]]
            rows_list.append(row_data)
        match = re.search('(\\d{2})(\\d{2})$', data_file.replace('.xlsx', ''))
        month = int(match.group(1))
        day = int(match.group(2))
        date = f'{month}月{day}'
        xlsx_datas[date] = rows_list
    return xlsx_datas

@retry(stop=stop_after_attempt(5), wait=wait_random(2, 5))
def fetch_single_stock(prefix_stock):
    session = get_session()
    url = f'https://qt.gtimg.cn/q={prefix_stock}'
    try:
        res = session.get(url, timeout=10)
        content = res.text
        content = content.split('=')[1].strip('";')
        parts = content.split('~')
        price = parts[3]
        change_percent = parts[32]
        turnover_rate = parts[38]
        circulation_value = parts[44]
        name = parts[1]
        today_high = parts[33]
        today_low = parts[34]
        return {'code': prefix_stock, 'data': {'现价': price, '涨幅': change_percent, '换手率': turnover_rate, '流通市值': circulation_value, '名称': name, '今日最高价': today_high, '今日最低价': today_low}}
    except Exception as e:
        return None

async def fetch_single_stock_async(session, prefix_stock, semaphore):
    """异步获取单个股票数据"""
    url = f'https://qt.gtimg.cn/q={prefix_stock}'
    async with semaphore:
        for attempt in range(5):
            try:
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as response:
                    content = await response.text()
                    content = content.split('=')[1].strip('";')
                    parts = content.split('~')
                    return {'code': prefix_stock, 'data': {'现价': parts[3], '涨幅': parts[32], '换手率': parts[38], '流通市值': parts[44], '名称': parts[1], '今日最高价': parts[33], '今日最低价': parts[34]}}
            except Exception as e:
                if attempt == 4:
                    return
                await asyncio.sleep(2 + attempt)
                continue

async def fetch_stocks_batch_async(stock_list, batch_name='批次'):
    """批量异步获取股票数据"""
    if not stock_list:
        return []
    connector = TCPConnector(limit=800, limit_per_host=200)
    semaphore = asyncio.Semaphore(800)
    timeout = aiohttp.ClientTimeout(total=10, connect=5)
    async with ClientSession(connector=connector, timeout=timeout) as session:
        tasks = [fetch_single_stock_async(session, prefix_stock, semaphore) for prefix_stock in stock_list]
        print(f'[异步爬取] {batch_name}: 开始爬取 {len(stock_list)} 个股票...')
        results = await asyncio.gather(*tasks, return_exceptions=True)
        success_count = sum((1 for r in results if r is not None and (not isinstance(r, Exception))))
        print(f'[异步爬取] {batch_name}: 完成，成功 {success_count}/{len(stock_list)} 个')
        return results

def get_real_time_data(progress_callback=None, strat_index=3, count=20, show_progress=True, top_priority_codes=None, high_priority_codes=None):
    """获取实时数据（异步版本，支持三级优先级）
    Args:
        top_priority_codes: 最高优先级（表格显示的股票）
        high_priority_codes: 高优先级（阳天数=1且有连续涨停）
    """
    start_time = time.time()
    prefix_stocks = []
    stock_dates = {}
    all_data = get_folder_data(strat_index=strat_index, count=count)
    all_stock_codes = []
    for date, stocks_list in all_data.items():
        for stock_data in stocks_list:
            stock_code = stock_data[0]
            all_stock_codes.append(stock_code)
    unique_stock_codes = list(set(all_stock_codes))
    total_count = len(unique_stock_codes)
    if progress_callback and show_progress:
        progress_callback(0, total_count, '开始爬取实时数据...')
    for stock in unique_stock_codes:
        if stock.startswith(('6', '9')):
            prefix_stocks.append('sh' + stock)
        elif stock.startswith(('0', '2', '3')):
            prefix_stocks.append('sz' + stock)
    top_priority_stocks = []
    high_priority_stocks = []
    normal_priority_stocks = []
    if top_priority_codes and len(top_priority_codes) > 0:
        top_priority_set = set()
        for code in top_priority_codes:
            if code.startswith(('6', '9')):
                top_priority_set.add('sh' + code)
            elif code.startswith(('0', '2', '3')):
                top_priority_set.add('sz' + code)
        for stock in prefix_stocks:
            if stock in top_priority_set:
                top_priority_stocks.append(stock)
        print(f'\n【最高优先级】表格显示股票: {len(top_priority_stocks)}个')
    if high_priority_codes and len(high_priority_codes) > 0:
        high_priority_set = set()
        for code in high_priority_codes:
            if code.startswith(('6', '9')):
                high_priority_set.add('sh' + code)
            elif code.startswith(('0', '2', '3')):
                high_priority_set.add('sz' + code)
        for stock in prefix_stocks:
            if stock not in set(top_priority_stocks) and stock in high_priority_set:
                high_priority_stocks.append(stock)
        print(f'【高优先级】阳天数=1且有连续涨停: {len(high_priority_stocks)}个')
    for stock in prefix_stocks:
        if stock not in set(top_priority_stocks) and stock not in set(high_priority_stocks):
            normal_priority_stocks.append(stock)
    print(f'【普通优先级】其他股票: {len(normal_priority_stocks)}个')
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        if top_priority_stocks:
            top_results = loop.run_until_complete(fetch_stocks_batch_async(top_priority_stocks, '最高优先级'))
            for result in top_results:
                if result and (not isinstance(result, Exception)):
                    stock_dates[result['code']] = result['data']
        if high_priority_stocks:
            high_results = loop.run_until_complete(fetch_stocks_batch_async(high_priority_stocks, '高优先级'))
            for result in high_results:
                if result and (not isinstance(result, Exception)):
                    stock_dates[result['code']] = result['data']
        if normal_priority_stocks:
            normal_results = loop.run_until_complete(fetch_stocks_batch_async(normal_priority_stocks, '普通优先级'))
            for result in normal_results:
                if result and (not isinstance(result, Exception)):
                    stock_dates[result['code']] = result['data']
    finally:
        loop.close()
    print(f'\n最终成功获取{len(stock_dates)}个实时数据')
    failed_count = total_count - len(stock_dates)
    if failed_count > 0:
        print(f'最终失败{failed_count}个实时数据')
    end_time = time.time()
    used_time = end_time - start_time
    if progress_callback and show_progress:
        progress_callback(total_count, total_count, '实时数据爬取完成')
    print(f'实时数据获取完成，用时 {used_time:.2f} 秒，成功 {len(stock_dates)} 个')
    return stock_dates

@retry(stop=stop_after_attempt(3), wait=wait_random(1, 3))
def get_code_industry():
    industry_dict = {}
    path = get_data_path('Table(1).xls')
    with open(path, 'r', encoding='gbk') as f:
        lines = f.readlines()
        for line in lines[1:]:
            parts = line.strip().split('\t')
            if len(parts) < 3:
                continue
            code, name, industry = parts[:3]
            code = code.lower()
            # 去掉可能的前缀，只保留数字部分
            if code.startswith(('sz', 'sh')):
                code = code[2:]
            industry_dict[code] = {'名字': name, '行业': industry}
    return industry_dict

@retry(stop=stop_after_attempt(5), wait=wait_random(2, 5))
def fetch_history_single(stock_code):
    try:
        session = get_session()
        if stock_code.startswith(('0', '2', '3')):
            secid = 0
        else:
            secid = 1
        stock_code_with_prefix = f'{secid}.{stock_code}'
        url = f'https://push2his.eastmoney.com/api/qt/stock/fflow/daykline/get?lmt=0&klt=101&fields1=f1%2Cf2%2Cf3%2Cf7&fields2=f51%2Cf52%2Cf53%2Cf54%2Cf55%2Cf56%2Cf57%2Cf58%2Cf59%2Cf60%2Cf61%2Cf62%2Cf63%2Cf64%2Cf65&ut=b2884a393a59ad64002292a3e90d46a5&secid={stock_code_with_prefix}'
        res = session.get(url, timeout=10)
        content = res.text
        start = content.find('(')
        end = content.rfind(')')
        if start != -1 and end != -1:
            json_str = content[start + 1:end]
        else:
            json_str = content
        data = json.loads(json_str)
        if data.get('rc') != 0 or 'data' not in data:
            return
        klines = data['data'].get('klines', [])
        if not klines:
            return
        prices = []
        for kline in klines:
            parts = kline.split(',')
            date = parts[0]
            close_price = float(parts[-4])
            change_pct = float(parts[-3])
            prices.append({'日期': date, '收盘价': close_price, '涨幅': change_pct})
        if len(prices) < 1:
            return
        if len(prices) > 61:
            prices = prices[-61:]
        prices_for_calc = prices[:-1]
        if len(prices_for_calc) < 1:
            return
        close_prices_for_calc = [p['收盘价'] for p in prices_for_calc]
        if len(close_prices_for_calc) >= 30:
            prices_30d = close_prices_for_calc[-30:]
        else:
            prices_30d = close_prices_for_calc
        max_30d = max(prices_30d)
        min_30d = min(prices_30d)
        if len(close_prices_for_calc) >= 60:
            prices_60d = close_prices_for_calc[-60:]
        else:
            prices_60d = close_prices_for_calc
        max_60d = max(prices_60d)
        min_60d = min(prices_60d)
        historical_prices = prices
        time_info = get_current_time_info()
        weekday = time_info['星期']
        hour = time_info['小时']
        minute = time_info['分钟']
        is_trading_time = weekday < 5 and (hour == 9 and minute >= 15 or 9 < hour < 15 or (hour == 15 and minute == 0))
        if is_trading_time:
            yesterday_close = prices[-1]['收盘价'] if len(prices) >= 1 else 0
            yesterday_change = prices[-1]['涨幅'] if len(prices) >= 1 else 0
        else:
            yesterday_close = prices[-2]['收盘价'] if len(prices) >= 2 else prices[-1]['收盘价'] if len(prices) >= 1 else 0
            yesterday_change = prices[-2]['涨幅'] if len(prices) >= 2 else prices[-1]['涨幅'] if len(prices) >= 1 else 0
        return {'代码': stock_code, '昨日收盘价': yesterday_close, '昨日涨幅': yesterday_change, '30日最高价': max_30d, '30日最低价': min_30d, '60日最高价': max_60d, '60日最低价': min_60d, '历史价格列表': historical_prices}
    except json.JSONDecodeError as e:
        return
    except Exception as e:
        return None

def get_history_data(progress_callback=None, strat_index=3, count=20, show_progress=True):
    """获取历史数据（优先从文件读取，不存在则爬取并保存）
    Args:
        progress_callback: 进度回调函数
        strat_index: 开始索引
        count: 文件数量
        show_progress: 是否显示进度
    Returns:
        dict: 历史数据字典
    """
    today_date = datetime.now().strftime('%Y-%m-%d')
    saved_data = load_history_data_from_file(today_date)
    if saved_data:
        from datetime import timedelta
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        if len(saved_data) > 0:
            sample_code = list(saved_data.keys())[0]
            price_list = saved_data[sample_code].get('历史价格列表', [])
            if price_list:
                latest_date = price_list[-1]['日期']
                if latest_date >= yesterday:
                    print(f'使用今天的历史数据文件 ({today_date}), 最新日期: {latest_date}')
                    if progress_callback and show_progress:
                        total_count = len(saved_data)
                        progress_callback(total_count, total_count, '从文件加载历史数据完成')
                    return saved_data
                print(f'历史数据文件过旧（最新日期: {latest_date}，应至少包含: {yesterday}），重新爬取...')
                try:
                    import os
                    folder_path = get_history_data_folder()
                    filename = get_history_data_filename(today_date)
                    file_path = os.path.join(folder_path, filename)
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        print(f'已删除过期缓存文件: {file_path}')
                except Exception as e:
                    print(f'删除过期缓存文件失败: {e}')
            else:
                print('历史数据文件无效（无价格数据），重新爬取...')
        else:
            print('历史数据文件为空，重新爬取...')
    print('今天的历史数据文件不存在，开始爬取...')
    all_data = get_folder_data(strat_index=strat_index, count=count)
    all_stock_codes = []
    stock_code_data = {}
    failed_stocks = []
    for date, stocks_list in all_data.items():
        for stock_data in stocks_list:
            stock_code = stock_data[0]
            all_stock_codes.append(stock_code)
    unique_stock_codes = list(set(all_stock_codes))
    total_count = len(unique_stock_codes)
    print(f'开始爬取 {total_count} 个股票的历史数据...')
    if progress_callback and show_progress:
        progress_callback(0, total_count, '开始爬取历史数据...')
    completed = [0]

    def fetch_with_progress(stock_code):
        try:
            result = fetch_history_single(stock_code)
            completed[0] += 1
            if progress_callback and show_progress:
                progress_callback(completed[0], total_count, f'历史数据: {completed[0]}/{total_count}')
            return (stock_code, result)
        except Exception as e:
            completed[0] += 1
            if progress_callback and show_progress:
                progress_callback(completed[0], total_count, f'历史数据: {completed[0]}/{total_count}')
            return (stock_code, None)
    with ThreadPoolExecutor(max_workers=250) as executor:
        results = executor.map(fetch_with_progress, unique_stock_codes)
    for stock_code, result in results:
        if result:
            stock_code_data[result['代码']] = result
        else:
            failed_stocks.append(stock_code)
    if failed_stocks:
        print(f'\n首次采集失败{len(failed_stocks)}个股票，开始二次重试...')
        retry_success = []
        retry_failed = []
        for i, stock_code in enumerate(failed_stocks):
            if progress_callback and show_progress:
                progress_callback(total_count + i + 1, total_count + len(failed_stocks), f'二次重试: {i + 1}/{len(failed_stocks)}')
            try:
                time.sleep(1)
                result = fetch_history_single(stock_code)
                if result:
                    stock_code_data[result['代码']] = result
                    retry_success.append(stock_code)
                    print(f'  ✓ {stock_code} 二次重试成功')
                else:
                    retry_failed.append(stock_code)
            except Exception as e:
                retry_failed.append(stock_code)
                print(f'  ✗ {stock_code} 二次重试仍失败: {e}')
        failed_stocks = retry_failed
        print(f'二次重试完成: 成功{len(retry_success)}个，仍失败{len(retry_failed)}个')
    print(f'\n最终成功获取{len(stock_code_data)}个股票数据')
    if failed_stocks:
        print(f'最终失败{len(failed_stocks)}个股票')
        save_failed_stocks(failed_stocks, today_date)
    if progress_callback and show_progress:
        progress_callback(total_count, total_count, '历史数据爬取完成')
    save_history_data_to_file(stock_code_data, today_date)
    return stock_code_data

def get_current_time_info():
    now = datetime.now()
    return {'小时': now.hour, '分钟': now.minute, '秒': now.second, '星期': now.weekday(), '时间': now.strftime('%H:%M:%S'), '日期': now.strftime('%Y-%m-%d')}

def should_use_yesterday_data():
    time_info = get_current_time_info()
    hour = time_info['小时']
    minute = time_info['分钟']
    if hour < 9 or (hour == 9 and minute < 15):
        return True
    return False

def get_data_source_index():
    time_info = get_current_time_info()
    weekday = time_info['星期']
    if weekday == 5:
        return (2, '周六使用周五数据')
    if weekday == 6:
        return (3, '周日使用周五数据')
    if should_use_yesterday_data():
        if weekday == 0:
            return (4, '周一9:15前使用周五数据')
        return (2, '9:15前使用昨天数据')
    return (1, '使用最新实时数据')

def check_data_updated(stock_code, old_price):
    try:
        if stock_code.startswith(('6', '9')):
            prefix_stock = 'sh' + stock_code
        else:
            prefix_stock = 'sz' + stock_code
        result = fetch_single_stock(prefix_stock)
        if result and result['data']:
            new_price = result['data'].get('现价', '')
            return (new_price != old_price, new_price)
        return (False, old_price)
    except:
        return (False, old_price)
if __name__ == '__main__':
    get_folder_data(3, 20)
    get_code_industry()
    get_history_data()